"""
Voylla Team Task Tracker
Department-based task board backed by Voylla Postgres.
- Each department logs in with its own password (voylla.task_tracker_departments)
- A task belongs to a department; only that department (or Admin) can edit it
- The department that raised a task can track it read-only
- Email notifications on assignment and status change (yagmail alert account)
Run:  python app.py   ->  http://<this-pc-ip>:5055
"""
import json
import os
import platform
import threading
import traceback
from datetime import date, datetime, timedelta

import psycopg2
import psycopg2.extras
from io import BytesIO

from flask import Flask, jsonify, request, render_template, session, redirect, send_file

# ---------------------------------------------------------------- credentials
# Priority: environment variables (cloud hosting) -> local credential files.
if platform.system() == "Windows":
    SCRIPTS_DIR = r"C:\Users\Amit Singh\Documents\Python_Scripts"
else:
    SCRIPTS_DIR = "/home/misauto/Python_Scripts"

APP_DIR = os.path.dirname(os.path.abspath(__file__))


def find_cred(fname):
    """Look next to the app first (creds/ folder), then the usual scripts dir."""
    for c in (os.path.join(APP_DIR, "creds", fname), os.path.join(SCRIPTS_DIR, fname)):
        if os.path.exists(c):
            return c
    raise FileNotFoundError(f"credential file not found: {fname}")


if os.environ.get("DB_HOST"):
    Voylla_config = {
        "host": os.environ["DB_HOST"],
        "user": os.environ["DB_USER"],
        "password": os.environ["DB_PASSWORD"],
        "database": os.environ["DB_NAME"],
        "port": os.environ.get("DB_PORT", "5432"),
    }
else:
    with open(find_cred("Voylla_Cred.txt")) as f:
        _lines = [l.strip() for l in f.readlines()]
    Voylla_config = {
        "host": _lines[0],
        "user": _lines[1],
        "password": _lines[2],
        "database": _lines[3],
        "port": _lines[4],
    }

# alert-email account (yagmail): email, app-password, sender name
if os.environ.get("MAIL_USER"):
    MAIL_USER = os.environ["MAIL_USER"]
    MAIL_PASS = os.environ["MAIL_PASS"]
    MAIL_SENDER = os.environ.get("MAIL_SENDER", "Automation Alert")
else:
    try:
        with open(find_cred("Automationalert_emailid_pass.txt")) as f:
            _m = [l.strip() for l in f.readlines()]
        MAIL_USER, MAIL_PASS, MAIL_SENDER = _m[0], _m[1], _m[2]
    except Exception:
        MAIL_USER = MAIL_PASS = MAIL_SENDER = None

TABLE = 'voylla.team_task_tracker'
DEPT_TABLE = 'voylla.task_tracker_departments'
COMMENT_TABLE = 'voylla.task_tracker_comments'
OUTBOX_TABLE = 'voylla.task_tracker_outbox'

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.secret_key = os.environ.get("SECRET_KEY") or ("voylla-task-tracker-" + Voylla_config["password"][:8])
app.permanent_session_lifetime = timedelta(days=90)


def get_conn():
    return psycopg2.connect(**Voylla_config)


def init_db():
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {TABLE} (
                id            SERIAL PRIMARY KEY,
                title         TEXT NOT NULL,
                requested_by  TEXT NOT NULL,
                priority      TEXT NOT NULL DEFAULT 'P3',
                their_tat     DATE,
                my_tat        DATE,
                status        TEXT NOT NULL DEFAULT 'Open',
                notes         TEXT,
                created_at    TIMESTAMPTZ NOT NULL DEFAULT now(),
                updated_at    TIMESTAMPTZ NOT NULL DEFAULT now(),
                completed_at  TIMESTAMPTZ
            );
        """)
        # new columns for department workflow (safe to re-run)
        cur.execute(f"""
            ALTER TABLE {TABLE}
                ADD COLUMN IF NOT EXISTS department TEXT,
                ADD COLUMN IF NOT EXISTS assigned_to TEXT,
                ADD COLUMN IF NOT EXISTS assignee_email TEXT,
                ADD COLUMN IF NOT EXISTS created_by_dept TEXT,
                ADD COLUMN IF NOT EXISTS requester_email TEXT,
                ADD COLUMN IF NOT EXISTS tat_status TEXT NOT NULL DEFAULT 'Pending',
                ADD COLUMN IF NOT EXISTS rating INTEGER,
                ADD COLUMN IF NOT EXISTS rating_comment TEXT,
                ADD COLUMN IF NOT EXISTS last_reminded DATE;
        """)
        cur.execute(f"UPDATE {TABLE} SET department = 'Admin' WHERE department IS NULL")
        cur.execute(f"UPDATE {TABLE} SET created_by_dept = department WHERE created_by_dept IS NULL")
        cur.execute(f"UPDATE {TABLE} SET tat_status = 'Accepted' WHERE my_tat IS NOT NULL AND tat_status = 'Pending'")

        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {OUTBOX_TABLE} (
                id         SERIAL PRIMARY KEY,
                recipients TEXT,
                subject    TEXT NOT NULL,
                html       TEXT NOT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                claimed_at TIMESTAMPTZ,
                sent_at    TIMESTAMPTZ,
                attempts   INTEGER NOT NULL DEFAULT 0,
                last_error TEXT
            );
        """)

        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {COMMENT_TABLE} (
                id         SERIAL PRIMARY KEY,
                task_id    INTEGER NOT NULL,
                dept       TEXT NOT NULL,
                author     TEXT,
                body       TEXT NOT NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT now()
            );
        """)

        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {DEPT_TABLE} (
                name     TEXT PRIMARY KEY,
                password TEXT NOT NULL
            );
        """)
        cur.execute(f"ALTER TABLE {DEPT_TABLE} ADD COLUMN IF NOT EXISTS email TEXT")
        # departments live in the DB; seed an Admin login only on a brand-new install
        cur.execute(f"SELECT COUNT(*) FROM {DEPT_TABLE}")
        if cur.fetchone()[0] == 0:
            import uuid
            admin_pw = os.environ.get("ADMIN_PASSWORD") or uuid.uuid4().hex[:10]
            cur.execute(f"INSERT INTO {DEPT_TABLE} (name, password) VALUES ('Admin', %s)", (admin_pw,))
            print(f"[init] created Admin department (password: {admin_pw})", flush=True)


def dept_list():
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(f"SELECT name FROM {DEPT_TABLE} ORDER BY name")
        return [r[0] for r in cur.fetchall()]


def row_to_dict(row):
    d = dict(row)
    for k in ("their_tat", "my_tat", "created_at", "updated_at", "completed_at"):
        if d.get(k):
            d[k] = d[k].isoformat()
    return d


# ---------------------------------------------------------------- email
BCC = os.environ.get("MAIL_BCC", "amit.singh@voylla.com")


def smtp_send(recipients, subject, html):
    """Actually send (raises on failure). recipients: list or None (BCC-only mail)."""
    import yagmail
    yag = yagmail.SMTP({MAIL_USER: MAIL_SENDER}, MAIL_PASS)
    yag.send(to=recipients or BCC, bcc=BCC if recipients else None, subject=subject, contents=html)


def _try_send_outbox(outbox_id):
    """Attempt immediate delivery of one queued mail; the mail worker retries later on failure."""
    try:
        with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                f"""UPDATE {OUTBOX_TABLE} SET claimed_at = now(), attempts = attempts + 1
                    WHERE id = %s AND sent_at IS NULL RETURNING *""",
                (outbox_id,),
            )
            row = cur.fetchone()
        if not row:
            return
        recips = json.loads(row["recipients"]) if row["recipients"] else None
        smtp_send(recips, row["subject"], row["html"])
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute(f"UPDATE {OUTBOX_TABLE} SET sent_at = now() WHERE id = %s", (outbox_id,))
        print(f"[mail] sent #{outbox_id} to {recips}", flush=True)
    except Exception as e:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute(f"UPDATE {OUTBOX_TABLE} SET last_error = %s, claimed_at = NULL WHERE id = %s",
                        (str(e)[:500], outbox_id))
        print(f"[mail] queued #{outbox_id} for worker (direct send failed: {e})", flush=True)


def notify(to, subject, html):
    """Queue an email in the outbox and try to send it immediately.
    If SMTP is blocked here (e.g. Render), the mail worker on the office
    server delivers it within a minute. Every mail BCCs MAIL_BCC."""
    if not (MAIL_USER and MAIL_PASS):
        return
    recips = [t for t in (to if isinstance(to, list) else [to]) if t and "@" in t]
    recips = list(dict.fromkeys(recips))
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            f"INSERT INTO {OUTBOX_TABLE} (recipients, subject, html) VALUES (%s, %s, %s) RETURNING id",
            (json.dumps(recips) if recips else None, subject, html),
        )
        oid = cur.fetchone()[0]
    threading.Thread(target=_try_send_outbox, args=(oid,), daemon=True).start()


def dept_email(name):
    if not name:
        return None
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(f"SELECT email FROM {DEPT_TABLE} WHERE name = %s", (name,))
        r = cur.fetchone()
    return r[0] if r and r[0] else None


PRI_STYLE = {"P1": ("#FEF2F2", "#DC2626", "P1 · Urgent"), "P2": ("#FFFBEB", "#B45309", "P2 · High"),
             "P3": ("#EFF6FF", "#1D4ED8", "P3 · Normal"), "P4": ("#F1F5F9", "#475569", "P4 · Low")}
ST_STYLE = {"Open": ("#EFF6FF", "#1D4ED8"), "In Progress": ("#EEF2FF", "#4338CA"), "Done": ("#F0FDF4", "#15803D")}


def _chip(text, bg, fg):
    return (f'<span style="display:inline-block;padding:4px 14px;border-radius:20px;background:{bg};'
            f'color:{fg};font-size:12px;font-weight:bold;font-family:Arial,sans-serif">{text}</span>')


def _fmt_d(v):
    if not v:
        return ""
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").strftime("%d %b %Y")
    except Exception:
        return str(v)


def task_mail_html(t, heading, base_url):
    def row(label, val):
        if not val:
            return ""
        return (f'<tr><td style="padding:10px 0;border-bottom:1px solid #F1F5F9;font-size:11px;color:#94A3B8;'
                f'text-transform:uppercase;letter-spacing:.6px;width:130px;vertical-align:middle;font-family:Arial,sans-serif">{label}</td>'
                f'<td style="padding:10px 0;border-bottom:1px solid #F1F5F9;font-size:14px;color:#0F172A;'
                f'font-family:Arial,sans-serif">{val}</td></tr>')

    pri = t.get("priority") or "P3"
    pb, pf, plabel = PRI_STYLE.get(pri, PRI_STYLE["P3"])
    st = t.get("status") or "Open"
    sb, sf = ST_STYLE.get(st, ST_STYLE["Open"])
    tat = t.get("tat_status") or "Pending"
    tb, tf = {"Pending": ("#FFFBEB", "#B45309"), "Accepted": ("#F0FDF4", "#15803D"),
              "Extended": ("#EFF6FF", "#1D4ED8")}.get(tat, ("#F1F5F9", "#475569"))
    rating = ""
    if t.get("rating"):
        rating = "★" * int(t["rating"]) + "☆" * (5 - int(t["rating"])) + f'&nbsp; {t["rating"]}/5'

    return f"""
<div style="background:#F4F6FA;padding:32px 12px;font-family:Arial,Helvetica,sans-serif">
  <table role="presentation" cellpadding="0" cellspacing="0" style="max-width:560px;margin:0 auto;width:100%">
    <tr><td style="background:#4F46E5;border-radius:14px 14px 0 0;padding:20px 28px">
      <table role="presentation" cellpadding="0" cellspacing="0" width="100%"><tr>
        <td style="color:#FFFFFF;font-size:17px;font-weight:bold;font-family:Arial,sans-serif">✓&nbsp; Voylla Task Tracker</td>
        <td align="right" style="color:#C7D2FE;font-size:12px;font-family:Arial,sans-serif">Task #{t.get('id','')}</td>
      </tr></table>
    </td></tr>
    <tr><td style="background:#FFFFFF;padding:28px;border:1px solid #E2E8F0;border-top:none;border-radius:0 0 14px 14px">
      <p style="margin:0 0 6px;font-size:13px;color:#4F46E5;font-weight:bold;text-transform:uppercase;letter-spacing:.5px">{heading}</p>
      <p style="margin:0 0 18px;font-size:19px;line-height:1.4;color:#0F172A;font-weight:bold">{t.get('title','')}</p>
      <p style="margin:0 0 22px">{_chip(plabel, pb, pf)}&nbsp;{_chip(st, sb, sf)}&nbsp;{_chip('TAT ' + tat, tb, tf)}</p>
      <table role="presentation" cellpadding="0" cellspacing="0" width="100%" style="margin-bottom:24px">
        {row('Department', t.get('department'))}
        {row('Assigned to', t.get('assigned_to'))}
        {row('Requested by', (t.get('requested_by') or '') + (' &nbsp;<span style="color:#94A3B8">(' + t.get('created_by_dept') + ')</span>' if t.get('created_by_dept') else ''))}
        {row('Needed by', _fmt_d(t.get('their_tat')))}
        {row('Committed date', _fmt_d(t.get('my_tat')))}
        {row('Rating', f'<span style="color:#B45309;font-size:15px">{rating}</span>' if rating else '')}
        {row('Notes', t.get('notes'))}
      </table>
      <table role="presentation" cellpadding="0" cellspacing="0"><tr><td style="border-radius:9px;background:#4F46E5">
        <a href="{base_url}" style="display:inline-block;padding:12px 28px;font-size:14px;font-weight:bold;color:#FFFFFF;text-decoration:none;font-family:Arial,sans-serif">Open Task Tracker&nbsp; →</a>
      </td></tr></table>
    </td></tr>
    <tr><td style="padding:18px 10px;text-align:center">
      <p style="margin:0;font-size:11.5px;color:#94A3B8;line-height:1.6">Automated notification from Voylla Task Tracker.<br>
      Please don't reply to this email — reply inside the task's conversation.</p>
    </td></tr>
  </table>
</div>"""


# ---------------------------------------------------------------- auth
def me():
    return session.get("dept")


def is_admin():
    return me() == "Admin"


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        dept = request.form.get("dept", "").strip()
        pw = request.form.get("key", "").strip()
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute(f"SELECT password FROM {DEPT_TABLE} WHERE name = %s", (dept,))
            row = cur.fetchone()
        if row and row[0] == pw:
            session.permanent = True
            session["dept"] = dept
            return redirect("/")
        error = "Wrong department password, try again."
    return render_template("login.html", departments=dept_list(), error=error), (401 if error else 200)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.before_request
def require_login():
    if request.path.startswith(("/login", "/logout", "/static", "/api/cron/")):
        return None  # cron endpoints authenticate with their own key
    if me():
        return None
    if request.path.startswith("/api/"):
        return jsonify({"error": "not logged in — refresh the page"}), 401
    return redirect("/login")


# ---------------------------------------------------------------- pages
@app.route("/")
def index():
    return render_template("index.html", dept=me(), admin=is_admin(), departments=dept_list())


# ---------------------------------------------------------------- api
@app.route("/api/tasks", methods=["GET"])
def list_tasks():
    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        if is_admin():
            cur.execute(f"SELECT * FROM {TABLE} ORDER BY id DESC")
        else:
            cur.execute(
                f"SELECT * FROM {TABLE} WHERE department = %s OR created_by_dept = %s ORDER BY id DESC",
                (me(), me()),
            )
        rows = cur.fetchall()
    out = []
    for r in rows:
        d = row_to_dict(r)
        d["can_edit"] = is_admin() or r["department"] == me()
        out.append(d)
    return jsonify(out)


@app.route("/api/tasks", methods=["POST"])
def create_task():
    data = request.get_json(force=True)
    title = (data.get("title") or "").strip()
    requested_by = (data.get("requested_by") or "").strip()
    department = (data.get("department") or "").strip()
    if not title or not requested_by:
        return jsonify({"error": "title and requested_by are required"}), 400
    if department not in dept_list():
        return jsonify({"error": "please select a valid department"}), 400
    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        # the giver only sets their_tat (needed-by); the receiving dept
        # accepts or extends it later, which is what fills my_tat
        cur.execute(
            f"""INSERT INTO {TABLE}
                (title, requested_by, priority, their_tat, my_tat, status, notes,
                 department, assigned_to, assignee_email, created_by_dept, requester_email, tat_status)
                VALUES (%s,%s,%s,%s,NULL,%s,%s,%s,%s,%s,%s,%s,'Pending') RETURNING *""",
            (
                title,
                requested_by,
                data.get("priority") or "P3",
                data.get("their_tat") or None,
                data.get("status") or "Open",
                (data.get("notes") or "").strip() or None,
                department,
                (data.get("assigned_to") or "").strip() or None,
                (data.get("assignee_email") or "").strip() or None,
                me(),
                (data.get("requester_email") or "").strip() or None,
            ),
        )
        row = cur.fetchone()
    t = row_to_dict(row)
    notify(
        [t.get("assignee_email"), dept_email(t["department"])],
        f"[Task Tracker] New {t['priority']} task for {t['department']}: {t['title']}",
        task_mail_html(t, f"{t['requested_by']} ({me()}) added a task for your department", request.host_url),
    )
    t["can_edit"] = is_admin() or t["department"] == me()
    return jsonify(t), 201


# their_tat / my_tat are deliberately NOT editable here — the committed date
# only moves through the accept/extend flow below
ALLOWED_FIELDS = {"title", "requested_by", "priority",
                  "status", "notes", "department", "assigned_to", "assignee_email", "requester_email"}


@app.route("/api/tasks/<int:task_id>", methods=["PATCH"])
def update_task(task_id):
    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(f"SELECT * FROM {TABLE} WHERE id = %s", (task_id,))
        existing = cur.fetchone()
    if not existing:
        return jsonify({"error": "not found"}), 404
    if not (is_admin() or existing["department"] == me()):
        return jsonify({"error": "only the " + (existing["department"] or "owner") + " department can change this task"}), 403

    data = request.get_json(force=True)
    fields = {k: v for k, v in data.items() if k in ALLOWED_FIELDS}
    if not fields:
        return jsonify({"error": "no valid fields"}), 400
    if "department" in fields and fields["department"] not in dept_list():
        return jsonify({"error": "invalid department"}), 400
    for k in ("their_tat", "my_tat"):
        if k in fields and not fields[k]:
            fields[k] = None

    sets = [f"{k} = %s" for k in fields]
    values = list(fields.values())
    sets.append("updated_at = now()")
    if fields.get("status") == "Done":
        sets.append("completed_at = COALESCE(completed_at, now())")
    elif "status" in fields:
        sets.append("completed_at = NULL")

    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"UPDATE {TABLE} SET {', '.join(sets)} WHERE id = %s RETURNING *",
            values + [task_id],
        )
        row = cur.fetchone()
    t = row_to_dict(row)
    if "status" in fields and fields["status"] != existing["status"]:
        notify(
            [t.get("assignee_email"), t.get("requester_email"),
             dept_email(t["department"]), dept_email(t.get("created_by_dept"))],
            f"[Task Tracker] {t['status']} — {t['title']}",
            task_mail_html(t, f"Task status changed to {t['status']}", request.host_url),
        )
    t["can_edit"] = is_admin() or t["department"] == me()
    return jsonify(t)


def get_task(task_id):
    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(f"SELECT * FROM {TABLE} WHERE id = %s", (task_id,))
        return cur.fetchone()


def can_see(t):
    return is_admin() or t["department"] == me() or t["created_by_dept"] == me()


def add_comment(task_id, author, body):
    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"INSERT INTO {COMMENT_TABLE} (task_id, dept, author, body) VALUES (%s,%s,%s,%s) RETURNING *",
            (task_id, me(), (author or "").strip() or None, body),
        )
        return row_to_dict(cur.fetchone())


@app.route("/api/tasks/<int:task_id>/tat", methods=["POST"])
def respond_tat(task_id):
    """Receiving department accepts the requested TAT or extends it with a note."""
    existing = get_task(task_id)
    if not existing:
        return jsonify({"error": "not found"}), 404
    if not (is_admin() or existing["department"] == me()):
        return jsonify({"error": "only the " + (existing["department"] or "owner") + " department can respond to the TAT"}), 403

    data = request.get_json(force=True)
    action = data.get("action")
    note = (data.get("note") or "").strip()

    if action == "accept":
        if not existing["their_tat"]:
            return jsonify({"error": "no requested TAT on this task to accept"}), 400
        new_tat = existing["their_tat"]
        tat_status = "Accepted"
        comment_body = f"✔ Accepted the requested TAT ({new_tat.strftime('%d %b %Y')})" + (f" — {note}" if note else "")
        mail_head = "Your requested TAT was accepted"
    elif action == "extend":
        if not data.get("my_tat"):
            return jsonify({"error": "pick the new committed date"}), 400
        if not note:
            return jsonify({"error": "please add a note explaining the new date"}), 400
        new_tat = data["my_tat"]
        tat_status = "Extended"
        comment_body = f"⏱ Committed TAT set to {datetime.strptime(new_tat, '%Y-%m-%d').strftime('%d %b %Y')} — {note}"
        mail_head = "The TAT on your task was extended"
    else:
        return jsonify({"error": "action must be accept or extend"}), 400

    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"UPDATE {TABLE} SET my_tat = %s, tat_status = %s, updated_at = now() WHERE id = %s RETURNING *",
            (new_tat, tat_status, task_id),
        )
        row = cur.fetchone()
    add_comment(task_id, data.get("author"), comment_body)

    t = row_to_dict(row)
    notify(
        [t.get("requester_email"), t.get("assignee_email"), dept_email(t.get("created_by_dept"))],
        f"[Task Tracker] {mail_head}: {t['title']}",
        task_mail_html(t, mail_head, request.host_url),
    )
    t["can_edit"] = is_admin() or t["department"] == me()
    return jsonify(t)


@app.route("/api/tasks/<int:task_id>/comments", methods=["GET", "POST"])
def comments(task_id):
    existing = get_task(task_id)
    if not existing:
        return jsonify({"error": "not found"}), 404
    if not can_see(existing):
        return jsonify({"error": "you don't have access to this task"}), 403

    if request.method == "GET":
        with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(f"SELECT * FROM {COMMENT_TABLE} WHERE task_id = %s ORDER BY id", (task_id,))
            return jsonify([row_to_dict(r) for r in cur.fetchall()])

    data = request.get_json(force=True)
    body = (data.get("body") or "").strip()
    if not body:
        return jsonify({"error": "empty comment"}), 400
    c = add_comment(task_id, data.get("author"), body)

    # notify the other side of the conversation
    t = row_to_dict(existing)
    if me() == existing["department"]:
        others = [existing["requester_email"], dept_email(existing.get("created_by_dept"))]
    else:
        others = [existing["assignee_email"], dept_email(existing["department"])]
    who = (data.get("author") or me())
    notify(
        others,
        f"[Task Tracker] New comment on: {t['title']}",
        task_mail_html(t, f"{who} ({me()}) commented: “{body}”", request.host_url),
    )
    return jsonify(c), 201


@app.route("/api/tasks/<int:task_id>/rate", methods=["POST"])
def rate_task(task_id):
    """The requesting department rates the work once it's Done."""
    existing = get_task(task_id)
    if not existing:
        return jsonify({"error": "not found"}), 404
    if not (is_admin() or existing["created_by_dept"] == me()):
        return jsonify({"error": "only the requesting department can rate this task"}), 403
    if existing["status"] != "Done":
        return jsonify({"error": "you can rate only after the task is marked Done"}), 400

    data = request.get_json(force=True)
    try:
        rating = int(data.get("rating"))
        assert 1 <= rating <= 5
    except Exception:
        return jsonify({"error": "rating must be 1-5"}), 400
    comment = (data.get("comment") or "").strip()

    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"UPDATE {TABLE} SET rating = %s, rating_comment = %s, updated_at = now() WHERE id = %s RETURNING *",
            (rating, comment or None, task_id),
        )
        row = cur.fetchone()
    stars = "★" * rating + "☆" * (5 - rating)
    add_comment(task_id, data.get("author"), f"⭐ Rated {rating}/5 {stars}" + (f" — {comment}" if comment else ""))

    t = row_to_dict(row)
    notify(
        [t.get("assignee_email"), dept_email(t["department"])],
        f"[Task Tracker] Rated {rating}/5 — {t['title']}",
        task_mail_html(t, f"Your work was rated {rating}/5 {stars}" + (f": “{comment}”" if comment else ""), request.host_url),
    )
    t["can_edit"] = is_admin() or t["department"] == me()
    return jsonify(t)


@app.route("/api/export")
def export_tasks():
    """Excel export of tasks visible to the logged-in department, with optional filters."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    q = f"SELECT * FROM {TABLE} WHERE 1=1"
    params = []
    if not is_admin():
        q += " AND (department = %s OR created_by_dept = %s)"
        params += [me(), me()]
    if request.args.get("dept"):
        q += " AND department = %s"
        params.append(request.args["dept"])
    if request.args.get("status") and request.args["status"] not in ("all", "active"):
        q += " AND status = %s"
        params.append(request.args["status"])
    if request.args.get("status") == "active":
        q += " AND status != 'Done'"
    if request.args.get("pri"):
        q += " AND priority = %s"
        params.append(request.args["pri"])
    if request.args.get("from"):
        q += " AND created_at::date >= %s"
        params.append(request.args["from"])
    if request.args.get("to"):
        q += " AND created_at::date <= %s"
        params.append(request.args["to"])
    q += " ORDER BY id"

    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(q, params)
        rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    headers = ["ID", "Task", "Department", "Assigned To", "Requested By", "Raised By Dept",
               "Priority", "Status", "Needed By", "Committed", "TAT Status", "Rating",
               "Created", "Completed", "Notes"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F46E5")
    for r in rows:
        ws.append([
            r["id"], r["title"], r["department"], r["assigned_to"], r["requested_by"],
            r["created_by_dept"], r["priority"], r["status"],
            str(r["their_tat"] or ""), str(r["my_tat"] or ""), r["tat_status"],
            r["rating"], r["created_at"].strftime("%Y-%m-%d %H:%M"),
            r["completed_at"].strftime("%Y-%m-%d %H:%M") if r["completed_at"] else "",
            r["notes"],
        ])
    for i, w in enumerate([6, 45, 16, 15, 15, 16, 8, 11, 11, 11, 10, 7, 16, 16, 40], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"voylla_tasks_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/cron/reminders")
def cron_reminders():
    """Daily TAT reminders: due tomorrow / due today / overdue follow-ups.
    Called by a scheduler (GitHub Actions). Max one reminder per task per day."""
    if request.args.get("key") != os.environ.get("CRON_KEY", "voylla-cron-2026"):
        return jsonify({"error": "bad key"}), 403

    today = date.today()
    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(f"""
            SELECT * FROM {TABLE}
            WHERE status != 'Done'
              AND COALESCE(my_tat, their_tat) IS NOT NULL
              AND COALESCE(my_tat, their_tat) <= %s
              AND (last_reminded IS NULL OR last_reminded < %s)
        """, (today + timedelta(days=1), today))
        due = cur.fetchall()

    sent = 0
    for r in due:
        t = row_to_dict(r)
        ref = r["my_tat"] or r["their_tat"]
        if ref == today + timedelta(days=1):
            head, subj = "⏰ This task is due TOMORROW", "Due tomorrow"
        elif ref == today:
            head, subj = "🔔 This task is due TODAY", "Due TODAY"
        else:
            days = (today - ref).days
            head, subj = f"🔴 This task is OVERDUE by {days} day{'s' if days > 1 else ''} — please update it", f"OVERDUE {days}d — follow-up"
        notify(
            [t.get("assignee_email"), dept_email(t["department"]),
             t.get("requester_email"), dept_email(t.get("created_by_dept"))],
            f"[Task Tracker] {subj}: {t['title']}",
            task_mail_html(t, head, "https://voyllatasker.onrender.com/"),
        )
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute(f"UPDATE {TABLE} SET last_reminded = %s WHERE id = %s", (today, r["id"]))
        sent += 1
    return jsonify({"reminders_sent": sent, "date": today.isoformat()})


@app.route("/api/tasks/<int:task_id>", methods=["DELETE"])
def delete_task(task_id):
    with get_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(f"SELECT department FROM {TABLE} WHERE id = %s", (task_id,))
        existing = cur.fetchone()
        if not existing:
            return jsonify({"error": "not found"}), 404
        if not (is_admin() or existing["department"] == me()):
            return jsonify({"error": "only the " + (existing["department"] or "owner") + " department can delete this task"}), 403
        cur.execute(f"DELETE FROM {COMMENT_TABLE} WHERE task_id = %s", (task_id,))
        cur.execute(f"DELETE FROM {TABLE} WHERE id = %s", (task_id,))
    return jsonify({"ok": True})


# run migrations/seeding on import too (gunicorn on Render never runs __main__)
try:
    init_db()
except Exception:
    traceback.print_exc()

if __name__ == "__main__":
    print("Voylla Task Tracker running -> http://0.0.0.0:5055")
    app.run(host="0.0.0.0", port=5055, debug=False)
